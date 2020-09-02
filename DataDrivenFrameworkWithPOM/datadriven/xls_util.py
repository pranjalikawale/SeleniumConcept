import openpyxl

class XlsUtility:   
    #get sheet object
    @staticmethod
    def get_sheet_object(path,sheet_name):
        #load workbook 
        workbook=openpyxl.load_workbook(path)
        #return sheet object by name
        sheet_object=workbook.get_sheet_by_name(sheet_name)
        #return sheet object
        return sheet_object

    #get row count
    @staticmethod
    def get_row_count(path,sheet_name):
        #invoke get_sheet_object method
        sheet=XlsUtility.get_sheet_object(path,sheet_name)
        #return row count
        return sheet.max_row    

    #get column count
    @staticmethod
    def get_column_count(path,sheet_name):
        #invoke get_sheet_object method
        sheet=XlsUtility.get_sheet_object(path,sheet_name)
        #return col count
        return sheet.max_column

    #read data from xls
    @staticmethod
    def read_data(path,sheet_name,row_count,col_count):
        #invoke get_sheet_object method
        sheet=XlsUtility.get_sheet_object(path,sheet_name)
        #return data from cell
        return sheet.cell(row=row_count,column=col_count).value 

    #write data into xls
    @staticmethod
    def write_data(path,sheet_name,row_count,col_count,data):
        #load workbook 
        workbook=openpyxl.load_workbook(path)
        #return sheet object by name
        sheet=workbook.get_sheet_by_name(sheet_name)
        #write data into cell
        sheet.cell(row=row_count,column=col_count).value=data
        #save modification in sheet
        workbook.save(path)
